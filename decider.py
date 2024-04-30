import os
import json
import boto3
import constants
import config
from openpyxl import Workbook, load_workbook
from decimal import Decimal


def create_excel_file():
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = constants.GROUP_LIST
    ws1['A1'] = "Name"
    ws1['B1'] = "Favorite Coffee"
    ws1['C1'] = "Turn To Pay"
    
    ws2 = wb.create_sheet(title=constants.ABSENTEE_LIST)
    ws2['A1'] = "Name"

    ws3 = wb.create_sheet(title = constants.MENU)
    ws3['A1'] = "Coffee Name"
    ws3['B1'] = "Price"
    ws3['A2'] = "Black Coffee"
    ws3['B2'] = "2.50"
    ws3['A3'] = "Mocha"
    ws3['B3'] = "3.00"
    ws3['A4'] = "Frappuccino"
    ws3['B4'] = "4.00"

    wb.save(constants.FILENAME)
    return True
    
def open_or_create_excel():
    if os.path.exists(constants.FILENAME):
        wb = load_workbook(constants.FILENAME)
        return False
    return  create_excel_file()
       
def print_menu(menu_dict,header1,header2):
    max_key_length = max(len(str(key)) for key in menu_dict.keys())
    max_value_length = max(len(str(value)) for value in menu_dict.values())

    print()
    print(f"{header1:{max_key_length}}  {header2:{max_value_length}}")
    print("-" * (max_key_length + max_value_length + 2))

    for key, value in menu_dict.items():
        print(f"{key:{max_key_length}} : ${value:{max_value_length}}")
    print()


def create_group(menu_dict):
    coffee_group = []
    while True:
        user_input = input("How many people are in your coffee run on a regular basis?: ")

        if user_input.isdigit() and int(user_input) > 1:
            for i in range(int(user_input)):
                user_name = input(f"What is person number {i+1} name? ")
                print_menu(menu_dict,"Coffee","Price")
                valid_order = None
                while valid_order is None:
                    order = input(f"What is {user_name}'s favorite order? ")
                    if order in menu_dict:
                        valid_order = order
                    else:
                        print("Invalid order. Please enter a valid order.")
                user_selection = (user_name,Decimal(menu_dict.get(valid_order)))
                coffee_group.append(user_selection)
                save_names_fav_order(user_name,valid_order)
            update_turn_cell(coffee_group[0][0])
            break
        else:
            print("Invalid input! Please enter a non-negative integer greater than 1 because this is a group.")
    
    return coffee_group

def update_turn_cell(name):
    wb, ws = load_wb_ws(constants.GROUP_LIST)
    ws['C2'] = name
    next_turn =  ws['C2'].value
    wb.save(constants.FILENAME)
    wb.close()
    return next_turn

def save_names_fav_order(string1, string2):
    wb, ws = load_wb_ws(constants.GROUP_LIST)
    next_row_a = ws.max_row + 1
    ws.cell(row=next_row_a, column=1, value=string1)
    next_row_b = next_row_a
    ws.cell(row=next_row_b, column=2, value=string2)
    wb.save(constants.FILENAME)

def read_turn(turn_index,coffee_group,today_group):
    wb, ws = load_wb_ws(constants.GROUP_LIST)
    today_payer = ws["C2"].value
    wb.save(constants.FILENAME)
    turn_index = get_names(coffee_group).index(ws["C2"].value)
    absent_payer = get_absent_override()
    total = calculate_amt(today_group)
    if absent_payer in get_names(today_group):
        today_payer = absent_payer
        remove_absence()
        print(f"{today_payer} was absent when they needed to pay, so they will pay now. They owe ${total}.")
        return turn_index
    absence_recorded = False
    while True:
        try:
            get_names(today_group).index(ws["C2"].value)
            today_payer = ws["C2"].value
            break
        except:
            print(f"{today_payer} is not here today, someone else will pay today and {today_payer} will pay when they get back")
            if not absence_recorded:
                absence_recorded = record_absence(today_payer)
            turn_index = (turn_index + 1) % len(coffee_group)
            update_turn_cell(coffee_group[turn_index][0])
            wb, ws = load_wb_ws(constants.GROUP_LIST)
            today_payer = ws["C2"].value
            wb.save(constants.FILENAME)

    print(f"It is {today_payer}'s turn to pay and they owe  ${total}.")
    turn_index = (turn_index + 1) % len(coffee_group)
    update_turn_cell(coffee_group[turn_index][0])
    return turn_index

def calculate_amt(today_group):
    total = sum(value for _, value in today_group)
    return total

def get_coffee_group(menu_dict):
    wb, ws = load_wb_ws(constants.GROUP_LIST)
    names = [cell.value for cell in ws['A'][1:]]
    fav_coffee = [cell.value for cell in ws['B'][1:]]
    price_array = []
    for fav in fav_coffee:
        price = Decimal(menu_dict.get(fav))
        price_array.append(price)

    coffee_group = list(dict(zip(names, price_array)).items())

    return coffee_group

def get_names(coffee_group):
    names = []
    for key, _ in coffee_group:
        names.append(key)
    return names

def record_absence(name):
    wb, ws = load_wb_ws(constants.ABSENTEE_LIST)
    row_index = 1
    while ws.cell(row=row_index, column=1).value is not None:
        row_index += 1
    
    ws.cell(row=row_index, column=1, value=name)
    print(f"Recording Absence of {name}")
    wb.save(constants.FILENAME)
    wb.close()
    return True

def remove_absence():
    wb, ws = load_wb_ws(constants.ABSENTEE_LIST)
    ws.delete_rows(2)
    wb.save(constants.FILENAME)

def get_absent_override():
    wb, ws = load_wb_ws(constants.ABSENTEE_LIST)
    oldest_absence = ws['A2'].value
    wb.close()
    return oldest_absence

def today_group_creation(absentee_list, coffee_group):
    today_group = coffee_group[:]
    for name in absentee_list:
        today_group = [item for item in today_group if item[0] != name]
    return today_group

def get_menu_aws():   
    if not config.AWS_ACCESS_KEY_ID or not config.AWS_SECRET_ACCESS_KEY:
        print("AWS credentials not found in environment variables, will retrieve menu locally ")
        return get_menu()
    try:

        s3 = boto3.resource(
            service_name=config.SERVICE_NAME,
            region_name=config.AWS_REGION,
            aws_access_key_id=config.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=config.AWS_SECRET_ACCESS_KEY
        )
        s3_obj = s3.Bucket('my-coffee-bucket-1').Object('MenuExample.json').get()
        menu_json=json.loads(s3_obj['Body'].read().decode('utf-8'))
        coffee_menu_dict = {item['name']: item['price'] for item in menu_json}
    except Exception as e:
        print("An error occurred: while trying to retrieve menu from AWS, will retrieve locally")
        return get_menu()

    print("Used AWS to retrieve menu!")
    return coffee_menu_dict


def get_menu():
    coffee_menu_dict = {}
    wb, ws = load_wb_ws(constants.MENU)
    for row in ws.iter_rows(min_row=2, values_only=True):
        item, price = row[0], row[1]
        coffee_menu_dict[item] = price
    return coffee_menu_dict

def load_wb_ws(sheetname):
    wb = load_workbook(constants.FILENAME)
    ws = wb[sheetname]
    return wb, ws
